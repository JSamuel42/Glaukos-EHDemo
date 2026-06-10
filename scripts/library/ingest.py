"""Library ingestion: reads RRMM_Publications_v5.xlsx, emits articles.json,
filter-tree.json, and corpus.json into data/library/."""
import openpyxl
from datetime import datetime
from collections import defaultdict
import json
from pathlib import Path

# Run from repo root; resolve paths relative to it.
REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE = REPO_ROOT / 'data' / 'source' / 'RRMM_Publications_v5.xlsx'
OUT_DIR = REPO_ROOT / 'data' / 'library'
OUT_DIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(SOURCE, data_only=True)
ws = wb['Alnyx']

# Column index map (1-indexed, matching the v5 Excel)
COL = {
    'article_id': 1,
    'product_group': 2,
    'brand': 3,
    'inn': 4,
    'indication': 5,
    'title': 6,
    'authors': 7,
    'journal': 8,
    'pub_date': 9,
    'pub_type': 10,
    'study_type': 11,
    'study_design': 12,
    'geography': 13,
    'sponsor': 14,
    'population': 15,
    'interventions': 16,
    'outcomes': 17,
    'category_raw': 18,
    'subcategory_raw': 19,
    'scientific_narrative': 20,
    'value_message': 21,
    'objection_handler': 22,
    'abstract': 23,
    'pub_link': 24,
}


def cell_hyperlink(cell):
    return cell.hyperlink.target if cell.hyperlink else None


def parse_categories(cat_raw, sub_raw):
    """Returns list of {category, subcategories: [str, ...]} aligned by ;; positional groups."""
    if not cat_raw:
        return []
    cat_groups = [c.strip() for c in str(cat_raw).split(';;')]
    sub_groups = [s.strip() for s in str(sub_raw).split(';;')] if sub_raw else [''] * len(cat_groups)
    out = []
    for cg, sg in zip(cat_groups, sub_groups):
        if not cg:
            continue
        subs = [s.strip() for s in sg.split(',') if s.strip()] if sg else []
        out.append({'category': cg, 'subcategories': subs})
    return out


def product_display(brand, inn):
    """Combine Brand and INN as 'Brand (INN)'. If both equal (placeholders or CAR-T), show once."""
    if brand and inn:
        b, i = brand.strip(), inn.strip()
        if b == i:
            return b
        return f'{b} ({i})'
    return brand or inn or ''


# A. Read rows and build article records
articles = []
for r in range(2, ws.max_row + 1):
    # Skip empty rows (last few rows of a sheet sometimes have a phantom max_row)
    if not ws.cell(r, COL['article_id']).value:
        continue

    cat_raw = ws.cell(r, COL['category_raw']).value
    sub_raw = ws.cell(r, COL['subcategory_raw']).value
    pub_date = ws.cell(r, COL['pub_date']).value
    brand = ws.cell(r, COL['brand']).value
    inn = ws.cell(r, COL['inn']).value

    article_url = cell_hyperlink(ws.cell(r, COL['article_id'])) or ws.cell(r, COL['pub_link']).value

    article = {
        'id': ws.cell(r, COL['article_id']).value,
        'url': article_url,
        'product_group': ws.cell(r, COL['product_group']).value,
        'brand': brand,
        'inn': inn,
        'product_display': product_display(brand, inn),
        'indication': ws.cell(r, COL['indication']).value,
        'title': ws.cell(r, COL['title']).value,
        'authors': ws.cell(r, COL['authors']).value,
        'journal': ws.cell(r, COL['journal']).value,
        'pub_date': pub_date.isoformat() if isinstance(pub_date, datetime) else None,
        'pub_year': pub_date.year if isinstance(pub_date, datetime) else None,
        'pub_link': ws.cell(r, COL['pub_link']).value,
        'pub_type': ws.cell(r, COL['pub_type']).value,
        'study_type': ws.cell(r, COL['study_type']).value,
        'study_design': ws.cell(r, COL['study_design']).value,
        'geography': ws.cell(r, COL['geography']).value,
        'sponsor': ws.cell(r, COL['sponsor']).value,
        'population': ws.cell(r, COL['population']).value,
        'interventions': ws.cell(r, COL['interventions']).value,
        'outcomes': ws.cell(r, COL['outcomes']).value,
        'categories': parse_categories(cat_raw, sub_raw),
        'scientific_narrative_link': ws.cell(r, COL['scientific_narrative']).value,
        'value_message_link': ws.cell(r, COL['value_message']).value,
        'objection_handler_link': ws.cell(r, COL['objection_handler']).value,
        'abstract': ws.cell(r, COL['abstract']).value,
    }
    articles.append(article)

print(f'Read {len(articles)} articles')

# B. Build the filter tree
# Product Group -> set of product_display strings
product_tree = defaultdict(set)
for a in articles:
    if a['product_group']:
        product_tree[a['product_group']].add(a['product_display'])

product_filter = []
for group in sorted(product_tree.keys()):
    children = sorted(product_tree[group])
    product_filter.append({'group': group, 'children': children})

# Category -> Sub-category from Sheet1 controlled vocab
sheet1 = wb['Sheet1']
category_tree = defaultdict(set)
for r in range(2, sheet1.max_row + 1):
    cat = sheet1.cell(r, 1).value
    sub = sheet1.cell(r, 2).value
    if cat and sub:
        category_tree[cat.strip()].add(sub.strip())

category_filter = []
for cat in sorted(category_tree.keys()):
    subs = sorted(category_tree[cat])
    category_filter.append({'category': cat, 'subcategories': subs})


def distinct_sorted(field):
    vals = set()
    for a in articles:
        v = a.get(field)
        if v:
            vals.add(v)
    return sorted(vals)


filter_tree = {
    'products': product_filter,
    'indications': distinct_sorted('indication'),
    'pub_types': distinct_sorted('pub_type'),
    'study_types': distinct_sorted('study_type'),
    'geographies': distinct_sorted('geography'),
    'sponsors': distinct_sorted('sponsor'),
    'categories': category_filter,
}

# C. Build the corpus (keyed by Article ID, used by chat for grounding)
corpus = {}
for a in articles:
    parts = []
    parts.append(f"Title: {a['title']}")
    if a['authors']:
        parts.append(f"Authors: {a['authors']}")
    if a['journal']:
        year_str = f", {a['pub_year']}" if a['pub_year'] else ''
        parts.append(f"Source: {a['journal']}{year_str}")
    if a['study_design']:
        parts.append(f"Study design: {a['study_design']}")
    if a['population']:
        parts.append(f"Population: {a['population']}")
    if a['interventions']:
        parts.append(f"Interventions: {a['interventions']}")
    if a['outcomes']:
        parts.append(f"Key outcomes: {a['outcomes']}")
    if a['abstract']:
        parts.append(f"\nAbstract:\n{a['abstract']}")

    corpus[a['id']] = {
        'id': a['id'],
        'title': a['title'],
        'subtitle': f"{a['product_display']} · {a['indication'] or '—'} · {a['pub_year'] or '—'}",
        'text': '\n\n'.join(parts),
    }

# D. Write outputs
ARTICLES_PATH = OUT_DIR / 'articles.json'
ARTICLES_PATH.write_text(json.dumps(articles, indent=2, ensure_ascii=False), encoding='utf-8')
print(f'Wrote {ARTICLES_PATH} ({len(articles)} articles)')

FILTERS_PATH = OUT_DIR / 'filter-tree.json'
FILTERS_PATH.write_text(json.dumps(filter_tree, indent=2, ensure_ascii=False), encoding='utf-8')
print(f'Wrote {FILTERS_PATH}')

CORPUS_PATH = OUT_DIR / 'corpus.json'
CORPUS_PATH.write_text(json.dumps(corpus, indent=2, ensure_ascii=False), encoding='utf-8')
print(f'Wrote {CORPUS_PATH} ({len(corpus)} entries)')

# E. Quick stats summary
print('\n=== Summary ===')
print(f'Articles: {len(articles)}')
print(f'Product groups: {len(filter_tree["products"])}')
print(f'Indications: {filter_tree["indications"]}')
print(f'Pub types: {filter_tree["pub_types"]}')
print(f'Study types: {filter_tree["study_types"]}')
print(f'Sponsors: {filter_tree["sponsors"]}')
print(f'Geographies: {len(filter_tree["geographies"])}')
print(f'Categories: {len(filter_tree["categories"])}')
abstract_len_total = sum(len(a["abstract"] or "") for a in articles)
print(f'Total abstract characters: {abstract_len_total:,}')
print(f'Articles missing abstract: {sum(1 for a in articles if not a["abstract"])}')
print(f'Articles missing geography: {sum(1 for a in articles if not a["geography"])}')
