"""Document Hub ingestion: reads Docuhub_table.xlsx, normalises whitespace,
emits documents.json (for UI) and corpus.json (for chat retrieval)."""
import openpyxl
from datetime import datetime
from collections import defaultdict
import json
from pathlib import Path

# Resolve repo-root-relative paths regardless of CWD when invoked.
REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE = REPO_ROOT / 'data' / 'source' / 'Docuhub_table.xlsx'
OUT_DIR = REPO_ROOT / 'data' / 'dochub'
OUT_DIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(SOURCE, data_only=True)
ws = wb['Sheet1']

COL = {
    'date': 1, 'product': 2, 'geography': 3, 'document': 4,
    'description': 5, 'type': 6, 'tag': 7, 'agency': 8, 'summary': 9,
}


def s(v):
    """Trim whitespace from string cells; return None for blanks."""
    if v is None:
        return None
    if isinstance(v, str):
        out = v.strip()
        return out if out else None
    return v


# A. Read rows, filter empties, build records
documents = []
for r in range(2, ws.max_row + 1):
    doc_title = s(ws.cell(r, COL['document']).value)
    if not doc_title:
        continue  # skip blank rows

    raw_date = ws.cell(r, COL['date']).value
    date_iso = raw_date.isoformat() if isinstance(raw_date, datetime) else None

    doc = {
        'id': f'doc-{r:03d}',  # stable IDs from row position
        'date': date_iso,
        'product': s(ws.cell(r, COL['product']).value),
        'geography': s(ws.cell(r, COL['geography']).value),
        'title': doc_title,
        'description': s(ws.cell(r, COL['description']).value),
        'type': s(ws.cell(r, COL['type']).value),
        'tag': s(ws.cell(r, COL['tag']).value),
        'agency': s(ws.cell(r, COL['agency']).value),
        'summary': s(ws.cell(r, COL['summary']).value),  # may be None
        'has_summary': bool(s(ws.cell(r, COL['summary']).value)),
    }
    documents.append(doc)

print(f'Loaded {len(documents)} documents')

# B. Build the filter tree
def distinct_sorted(field, sort_key=None):
    vals = set()
    for d in documents:
        v = d.get(field)
        if v:
            vals.add(v)
    return sorted(vals, key=sort_key) if sort_key else sorted(vals)


# Geographies in a sensible UI order: Global first, then alphabetical
def geo_key(g):
    return (0 if g == 'Global' else 1, g)


filter_tree = {
    'products': distinct_sorted('product'),
    'geographies': distinct_sorted('geography', sort_key=geo_key),
    'types': distinct_sorted('type'),
    'tags': distinct_sorted('tag'),
}

print(f'Filters: {len(filter_tree["products"])} products, '
      f'{len(filter_tree["geographies"])} geographies, '
      f'{len(filter_tree["types"])} types, '
      f'{len(filter_tree["tags"])} tags')

# C. Build the chat corpus
corpus_entries = []
for d in documents:
    entry = {
        'id': d['id'],
        'product': d['product'],
        'geography': d['geography'],
        'title': d['title'],
        'description': d['description'] or '(No description)',
        'type': d['type'],
        'tag': d['tag'],
        'agency': d['agency'],
        'date': d['date'],
        'has_summary': d['has_summary'],
        'summary': d['summary'] if d['has_summary'] else None,
    }
    corpus_entries.append(entry)

# D. Build the table grouping structure
grouped = defaultdict(list)
for d in documents:
    grouped[(d['product'], d['geography'])].append(d)


def date_key(d):
    return d['date'] or '0000'


groups = []
for (product, geography), docs in grouped.items():
    docs_sorted = sorted(docs, key=date_key, reverse=True)
    groups.append({
        'product': product,
        'geography': geography,
        'doc_ids': [d['id'] for d in docs_sorted],
        'count': len(docs_sorted),
    })

# E. Write outputs
DOCS_PATH = OUT_DIR / 'documents.json'
DOCS_PATH.write_text(json.dumps({
    'documents': documents,
    'filter_tree': filter_tree,
    'groups': groups,
}, indent=2, ensure_ascii=False), encoding='utf-8')
print(f'Wrote {DOCS_PATH}: {len(documents)} documents, {len(groups)} groups')

CORPUS_PATH = OUT_DIR / 'corpus.json'
CORPUS_PATH.write_text(json.dumps({
    'entries': corpus_entries,
}, indent=2, ensure_ascii=False), encoding='utf-8')
print(f'Wrote {CORPUS_PATH}: {len(corpus_entries)} entries '
      f'({sum(1 for e in corpus_entries if e["has_summary"])} with summaries)')

# F. Quick stats
print('\n=== Summary ===')
print(f'Documents: {len(documents)}')
print(f'Documents with summary: {sum(1 for d in documents if d["has_summary"])}')
print(f'Product-country groups: {len(groups)}')
for p in filter_tree['products']:
    p_count = sum(1 for d in documents if d['product'] == p)
    print(f'  {p}: {p_count} docs')
print('\nGeographies:', filter_tree['geographies'])
print('Types:', filter_tree['types'])
print('Tags:', filter_tree['tags'])
